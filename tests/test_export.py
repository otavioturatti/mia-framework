"""Tests for export.py"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from mia.exceptions import MiaError
from mia.export import export_documents, read_identity, read_map, run
from mia.generate_template import (
    EXAMPLES,
    IDENTITY_FIELDS_START_ROW,
    IDENTITY_VALUE_COL,
    generate,
)


@pytest.fixture()
def template_path(tmp_path: Path) -> Path:
    """Generate a template in a temp directory."""
    output = tmp_path / "test_mia.xlsx"
    generate(str(output))
    return output


@pytest.fixture()
def filled_template(tmp_path: Path) -> Path:
    """Generate a template with identity fields filled in."""
    output = tmp_path / "filled_mia.xlsx"
    generate(str(output))

    wb = openpyxl.load_workbook(output)
    ws = wb["Identity"]
    col = IDENTITY_VALUE_COL
    row = IDENTITY_FIELDS_START_ROW
    ws.cell(row=row, column=col).value = "MIA Sales"
    ws.cell(row=row + 1, column=col).value = "Sales Module"
    ws.cell(row=row + 2, column=col).value = "PostgreSQL sales_db"
    ws.cell(row=row + 3, column=col).value = "sales-agent-v1"
    ws.cell(row=row + 4, column=col).value = "1.0"
    ws.cell(row=row + 5, column=col).value = "Covers sales questions"
    wb.save(output)
    wb.close()

    return output


class TestReadIdentity:
    def test_reads_empty_identity(self, template_path: Path) -> None:
        wb = openpyxl.load_workbook(template_path, read_only=True)
        identity = read_identity(wb["Identity"])
        wb.close()

        assert identity["name"] == ""
        assert identity["domain"] == ""
        assert len(identity) == 6

    def test_reads_filled_identity(self, filled_template: Path) -> None:
        wb = openpyxl.load_workbook(filled_template, read_only=True)
        identity = read_identity(wb["Identity"])
        wb.close()

        assert identity["name"] == "MIA Sales"
        assert identity["domain"] == "Sales Module"
        assert identity["data_source"] == "PostgreSQL sales_db"
        assert identity["agent"] == "sales-agent-v1"
        assert identity["version"] == "1.0"


class TestReadMap:
    def test_reads_all_examples(self, template_path: Path) -> None:
        wb = openpyxl.load_workbook(template_path, read_only=True)
        rows = read_map(wb["Map"])
        wb.close()

        assert len(rows) == len(EXAMPLES)

    def test_first_row_content(self, template_path: Path) -> None:
        wb = openpyxl.load_workbook(template_path, read_only=True)
        rows = read_map(wb["Map"])
        wb.close()

        first = rows[0]
        assert first["question"] == "How much did I sell yesterday?"
        assert "revenue" in first["intent"].lower()
        assert first["operation"] != ""
        assert first["action"].startswith("SQL:")

    def test_skips_empty_rows(self, template_path: Path) -> None:
        wb = openpyxl.load_workbook(template_path, read_only=True)
        rows = read_map(wb["Map"])
        wb.close()

        assert len(rows) == len(EXAMPLES)


class TestExportDocuments:
    def test_document_structure(self) -> None:
        identity = {"name": "Test MIA", "domain": "Test", "version": "1.0",
                     "data_source": "", "agent": "", "description": ""}
        rows = [{"question": "Q1", "intent": "I1", "operation": "O1", "action": "A1"}]

        docs = export_documents(identity, rows)

        assert len(docs) == 1
        doc = docs[0]
        assert doc["id"] == "test_mia_001"
        assert doc["content_for_embedding"] == "Q1 — I1"
        assert doc["content_for_reranking"] == "I1 — O1"
        assert doc["metadata"]["question"] == "Q1"
        assert doc["metadata"]["intent"] == "I1"
        assert doc["metadata"]["operation"] == "O1"
        assert doc["metadata"]["action"] == "A1"
        assert doc["metadata"]["domain"] == "Test"

    def test_embedding_format_is_question_separator_intent(self) -> None:
        """Verify exact format: '{question} — {intent}', not reversed or missing separator."""
        identity = {"name": "MIA", "domain": "", "version": "",
                     "data_source": "", "agent": "", "description": ""}
        rows = [{"question": "What is X?", "intent": "Find X value", "operation": "Lookup X", "action": "SQL"}]

        docs = export_documents(identity, rows)

        assert docs[0]["content_for_embedding"] == "What is X? — Find X value"
        assert docs[0]["content_for_reranking"] == "Find X value — Lookup X"

    def test_duplicate_questions_have_different_embeddings(self) -> None:
        """Duplicate questions with different intents must produce different embeddings."""
        identity = {"name": "MIA", "domain": "", "version": "",
                     "data_source": "", "agent": "", "description": ""}
        rows = [
            {"question": "How much did I sell?", "intent": "Store revenue", "operation": "Sum by store", "action": ""},
            {"question": "How much did I sell?", "intent": "Salesperson revenue", "operation": "Sum by user", "action": ""},
        ]

        docs = export_documents(identity, rows)

        assert docs[0]["content_for_embedding"] != docs[1]["content_for_embedding"]
        assert docs[0]["content_for_reranking"] != docs[1]["content_for_reranking"]

    def test_empty_intent_no_dangling_separator(self) -> None:
        """Question with empty intent should not have a trailing separator."""
        identity = {"name": "MIA", "domain": "", "version": "",
                     "data_source": "", "agent": "", "description": ""}
        rows = [{"question": "Q1", "intent": "", "operation": "", "action": "A1"}]

        docs = export_documents(identity, rows)

        assert docs[0]["content_for_embedding"] == "Q1"
        assert "—" not in docs[0]["content_for_embedding"]

    def test_empty_action_still_exported(self) -> None:
        """A row with question+intent+operation but empty action should still be exported."""
        identity = {"name": "MIA", "domain": "", "version": "",
                     "data_source": "", "agent": "", "description": ""}
        rows = [{"question": "Q1", "intent": "I1", "operation": "O1", "action": ""}]

        docs = export_documents(identity, rows)

        assert len(docs) == 1
        assert docs[0]["metadata"]["action"] == ""

    def test_unicode_and_accents(self) -> None:
        """Ensure non-ASCII characters are preserved correctly."""
        identity = {"name": "MIA Vendas", "domain": "Vendas", "version": "1.0",
                     "data_source": "", "agent": "", "description": ""}
        rows = [
            {"question": "Qual o faturamento de ontem?",
             "intent": "Faturamento total da loja",
             "operation": "Somatório das vendas",
             "action": "SQL: SELECT SUM(valor) FROM vendas"},
        ]

        docs = export_documents(identity, rows)

        assert "faturamento" in docs[0]["content_for_embedding"]
        assert "Somatório" in docs[0]["content_for_reranking"]
        assert docs[0]["metadata"]["question"] == "Qual o faturamento de ontem?"

    def test_sequential_ids(self) -> None:
        identity = {"name": "MIA", "domain": "", "version": "",
                     "data_source": "", "agent": "", "description": ""}
        rows = [
            {"question": f"Q{i}", "intent": "", "operation": "", "action": ""}
            for i in range(5)
        ]

        docs = export_documents(identity, rows)

        ids = [d["id"] for d in docs]
        assert ids == ["mia_001", "mia_002", "mia_003", "mia_004", "mia_005"]


class TestRunPipeline:
    def test_full_pipeline(self, filled_template: Path, tmp_path: Path) -> None:
        output = tmp_path / "output.json"
        result = run(filled_template, output)

        assert result == output
        assert output.exists()

        with open(output, encoding="utf-8") as f:
            docs = json.load(f)

        assert len(docs) == len(EXAMPLES)
        assert docs[0]["metadata"]["mia_name"] == "MIA Sales"
        assert docs[0]["metadata"]["domain"] == "Sales Module"
        assert "content_for_reranking" in docs[0]

    def test_default_output_name(self, filled_template: Path, monkeypatch: pytest.MonkeyPatch) -> None:
        monkeypatch.chdir(filled_template.parent)
        result = run(filled_template)

        expected = Path(filled_template.stem + "_export.json")
        assert result == expected
        assert expected.exists()

    def test_nonexistent_file_exits(self, tmp_path: Path) -> None:
        with pytest.raises(MiaError):
            run(tmp_path / "nonexistent.xlsx")

    def test_wrong_extension_exits(self, tmp_path: Path) -> None:
        bad_file = tmp_path / "bad.csv"
        bad_file.touch()
        with pytest.raises(MiaError):
            run(bad_file)

    def test_missing_identity_sheet_exits(self, tmp_path: Path) -> None:
        path = tmp_path / "no_identity.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Map"
        wb.save(path)
        wb.close()

        with pytest.raises(MiaError):
            run(path)

    def test_missing_map_sheet_exits(self, tmp_path: Path) -> None:
        path = tmp_path / "no_map.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Identity"
        wb.save(path)
        wb.close()

        with pytest.raises(MiaError):
            run(path)

    def test_corrupted_headers_raises(self, tmp_path: Path) -> None:
        """If Map sheet headers are modified, export should fail with a clear error."""
        from mia.generate_template import IDENTITY_FIELD_COL as ID_COL

        path = tmp_path / "bad_headers.xlsx"
        wb = openpyxl.Workbook()
        ws_id = wb.active
        ws_id.title = "Identity"
        # Write valid identity fields so identity validation passes
        fields = ["Name", "Domain", "Data Source", "Agent", "Version", "Description"]
        for i, f in enumerate(fields):
            ws_id.cell(row=IDENTITY_FIELDS_START_ROW + i, column=ID_COL).value = f
        ws_map = wb.create_sheet("Map")
        # Write wrong headers at the expected positions
        ws_map.cell(row=6, column=2).value = "Wrong"
        ws_map.cell(row=6, column=3).value = "Headers"
        ws_map.cell(row=6, column=4).value = "Here"
        ws_map.cell(row=6, column=5).value = "Oops"
        wb.save(path)
        wb.close()

        with pytest.raises(MiaError, match="headers don't match"):
            run(path)

    def test_corrupted_identity_fields_raises(self, tmp_path: Path) -> None:
        """If Identity sheet fields are modified, export should fail with a clear error."""
        path = tmp_path / "bad_identity.xlsx"
        wb = openpyxl.Workbook()
        ws_id = wb.active
        ws_id.title = "Identity"
        ws_id.cell(row=6, column=2).value = "Wrong Field"
        ws_map = wb.create_sheet("Map")
        ws_map.cell(row=6, column=2).value = "Question"
        ws_map.cell(row=6, column=3).value = "Intent"
        ws_map.cell(row=6, column=4).value = "Operation"
        ws_map.cell(row=6, column=5).value = "Action"
        ws_map.cell(row=8, column=2).value = "Test question"
        wb.save(path)
        wb.close()

        with pytest.raises(MiaError, match="Identity sheet fields don't match"):
            run(path)
