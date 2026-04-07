"""Tests for generate_template.py"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

from mia.exceptions import MiaError
from mia.generate_template import (
    EXAMPLES,
    IDENTITY_FIELDS,
    IDENTITY_FIELDS_START_ROW,
    IDENTITY_VALUE_COL,
    MAP_DATA_START_ROW,
    MAP_FIRST_COL,
    MAP_LAST_COL,
    generate,
)


@pytest.fixture()
def template_path(tmp_path: Path) -> Path:
    """Generate a template in a temp directory and return its path."""
    output = tmp_path / "test_template.xlsx"
    generate(str(output))
    return output


@pytest.fixture()
def workbook(template_path: Path) -> openpyxl.Workbook:
    wb = openpyxl.load_workbook(template_path)
    yield wb
    wb.close()


class TestTemplateGeneration:
    def test_file_is_created(self, template_path: Path) -> None:
        assert template_path.exists()
        assert template_path.stat().st_size > 0

    def test_has_both_sheets(self, workbook: openpyxl.Workbook) -> None:
        assert "Identity" in workbook.sheetnames
        assert "Map" in workbook.sheetnames

    def test_identity_sheet_has_title(self, workbook: openpyxl.Workbook) -> None:
        ws = workbook["Identity"]
        assert "MIA" in str(ws.cell(row=2, column=2).value)

    def test_identity_sheet_has_author(self, workbook: openpyxl.Workbook) -> None:
        ws = workbook["Identity"]
        assert "Joao Otavio" in str(ws.cell(row=4, column=2).value)

    def test_identity_fields_are_present(self, workbook: openpyxl.Workbook) -> None:
        ws = workbook["Identity"]
        for i, field in enumerate(IDENTITY_FIELDS):
            cell_value = ws.cell(row=IDENTITY_FIELDS_START_ROW + i, column=2).value
            assert cell_value == field

    def test_identity_values_are_empty(self, workbook: openpyxl.Workbook) -> None:
        ws = workbook["Identity"]
        for i in range(len(IDENTITY_FIELDS)):
            value = ws.cell(row=IDENTITY_FIELDS_START_ROW + i, column=IDENTITY_VALUE_COL).value
            assert value == "" or value is None

    def test_map_sheet_has_headers(self, workbook: openpyxl.Workbook) -> None:
        ws = workbook["Map"]
        headers = [ws.cell(row=6, column=col).value for col in range(MAP_FIRST_COL, MAP_LAST_COL + 1)]
        assert headers == ["Question", "Intent", "Operation", "Action"]

    def test_map_sheet_has_all_examples(self, workbook: openpyxl.Workbook) -> None:
        ws = workbook["Map"]
        for i, example in enumerate(EXAMPLES):
            row = MAP_DATA_START_ROW + i
            values = [ws.cell(row=row, column=col).value for col in range(MAP_FIRST_COL, MAP_LAST_COL + 1)]
            assert values == example

    def test_map_sheet_has_empty_rows(self, workbook: openpyxl.Workbook) -> None:
        ws = workbook["Map"]
        first_empty_row = MAP_DATA_START_ROW + len(EXAMPLES)
        values = [ws.cell(row=first_empty_row, column=col).value for col in range(MAP_FIRST_COL, MAP_LAST_COL + 1)]
        assert all(v is None for v in values)


class TestGenerateErrors:
    @pytest.mark.skipif(sys.platform == "win32", reason="chmod has no effect on Windows NTFS")
    def test_permission_error_raises(self, tmp_path: Path) -> None:
        output = tmp_path / "locked.xlsx"
        output.touch()
        output.chmod(0o000)
        with pytest.raises(MiaError):
            generate(str(output))
        output.chmod(0o644)
