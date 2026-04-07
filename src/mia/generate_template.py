"""
MIA Framework - Template Generator
Map of Intent and Action

Generates an .xlsx file with:
  - Sheet 1: Identity (MIA metadata)
  - Sheet 2: Map (core table: Question > Intent > Operation > Action)

Usage:
    python generate_template.py
    python generate_template.py --output my_mia.xlsx

Output:
    MIA_Template.xlsx (default)
"""

from __future__ import annotations

import argparse
import logging
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mia.exceptions import MiaError

# ---------------------------------------------------------------------------
# Layout constants — single source of truth for row/column positions
# ---------------------------------------------------------------------------

# Identity sheet
IDENTITY_TITLE_ROW = 2
IDENTITY_SUBTITLE_ROW = 3
IDENTITY_AUTHOR_ROW = 4
IDENTITY_FIELDS_START_ROW = 6
IDENTITY_FIELD_COL = 2  # Column B
IDENTITY_VALUE_COL = 3  # Column C

# Map sheet
MAP_TITLE_ROW = 2
MAP_SUBTITLE_ROW = 3
MAP_AUTHOR_ROW = 4
MAP_HEADER_ROW = 6
MAP_DESC_ROW = 7
MAP_DATA_START_ROW = 8
MAP_FIRST_COL = 2  # Column B
MAP_LAST_COL = 5   # Column E

EMPTY_ROWS_COUNT = 20

AUTHOR_LINE = (
    "Created with love by Joao Otavio Turatti Barbosa"
    "  |  linkedin.com/in/otavioturatti"
)

IDENTITY_FIELDS: list[str] = [
    "Name",
    "Domain",
    "Data Source",
    "Agent",
    "Version",
    "Description",
]

MAP_HEADERS: list[str] = ["Question", "Intent", "Operation", "Action"]
MAP_DESCRIPTIONS: list[str] = [
    "What did the user say?",
    "What did they mean?",
    "How to solve it?",
    "What to execute?",
]

EXAMPLES: list[list[str]] = [
    # SQL - Calculation
    [
        "How much did I sell yesterday?",
        "Total store revenue for the previous day",
        "Sum of sales values filtered by date = yesterday",
        "SQL: SELECT SUM(amount) FROM sales WHERE date = CURRENT_DATE - 1",
    ],
    # SQL - Same question, different profile
    [
        "How much did I sell yesterday?",
        "Individual salesperson revenue for the previous day",
        "Sum of sales values filtered by salesperson and date = yesterday",
        "SQL: SELECT SUM(amount) FROM sales WHERE salesperson_id = @user AND date = CURRENT_DATE - 1",
    ],
    # SQL - Direct lookup
    [
        "What is John's email?",
        "Find contact information for customer John Smith",
        "Filter by name in the customers table",
        "SQL: SELECT email FROM customers WHERE name ILIKE '%john smith%'",
    ],
    # SQL - Business jargon (customers = receipts)
    [
        "How many customers did I have yesterday?",
        "Number of receipts issued yesterday",
        "Count of receipts filtered by date = yesterday",
        "SQL: SELECT COUNT(*) FROM receipts WHERE date = CURRENT_DATE - 1",
    ],
    # REST API - External query
    [
        "What's the status of my order?",
        "Track current status of the customer's order",
        "Lookup order by user ID via logistics API",
        "API: GET /api/v1/tracking/{order_id} with header Authorization: Bearer @token",
    ],
    # Agent Tool - Notification
    [
        "Let the team know the report is ready",
        "Send notification to the team about report completion",
        "Identify team channel and compose notification message",
        "Tool: send_slack_message(channel='#sales-team', text='Report completed and available')",
    ],
    # API + Extraction
    [
        "What's the dollar exchange rate right now?",
        "Query current USD/EUR exchange rate",
        "Call exchange rate API and extract the sell value",
        "API: GET /api/v1/rates/USD-EUR -> return field 'sell'",
    ],
    # Orchestrator - Routing
    [
        "I need a report on sales and inventory",
        "Generate report crossing sales and inventory data",
        "Identify domains involved: Sales + Inventory",
        "Route -> Sales Agent (revenue data) + Inventory Agent (availability data)",
    ],
]

INSTRUCTIONS: list[str] = [
    "1. Fill in all fields above with your project data",
    "2. Go to the 'Map' sheet and fill in the intent rows",
    "3. Run 'mia-export <file>.xlsx' to generate the JSON for your pipeline",
    "",
    "Tips:",
    "- One MIA per domain in microservice architectures",
    "- One single MIA for monolithic projects",
    "- Duplicate rows in the Map when user profile changes the behavior",
    "- The Question should be representative (no typo variations - RAG handles it semantically)",
    "- Capture business jargon in the Intent (e.g., 'customers' = receipts issued in retail)",
]

logger = logging.getLogger("mia.generate")


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

def create_styles() -> dict[str, Any]:
    """Define the visual styles for the framework."""
    return {
        "title_font": Font(name="Calibri", size=20, bold=True, color="FFFFFF"),
        "subtitle_font": Font(name="Calibri", size=11, italic=True, color="B0B0B0"),
        "header_font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
        "field_font": Font(name="Calibri", size=11, bold=True, color="2D2D2D"),
        "value_font": Font(name="Calibri", size=11, color="4A4A4A"),
        "author_font": Font(name="Calibri", size=10, bold=False, color="6C8EBF"),
        "instruction_title_font": Font(name="Calibri", size=13, bold=True, color="1B1B2F"),
        "instruction_font": Font(name="Calibri", size=10, color="5A5A5A"),
        "desc_font": Font(name="Calibri", size=9, italic=True, color="FFFFFF"),
        "fill_dark": PatternFill(start_color="1B1B2F", end_color="1B1B2F", fill_type="solid"),
        "fill_header": PatternFill(start_color="162447", end_color="162447", fill_type="solid"),
        "fill_desc": PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid"),
        "fill_row_even": PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid"),
        "fill_row_odd": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        "fill_field": PatternFill(start_color="E8EDF3", end_color="E8EDF3", fill_type="solid"),
        "fill_value": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        "border": Border(
            left=Side(style="thin", color="D0D5DD"),
            right=Side(style="thin", color="D0D5DD"),
            top=Side(style="thin", color="D0D5DD"),
            bottom=Side(style="thin", color="D0D5DD"),
        ),
        "align_center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "align_left": Alignment(horizontal="left", vertical="center", wrap_text=True),
    }


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _fill_row_dark(ws: Worksheet, row: int, styles: dict[str, Any], col_range: range) -> None:
    """Apply dark background to a range of columns in a row."""
    for col in col_range:
        ws.cell(row=row, column=col).fill = styles["fill_dark"]


def create_identity_sheet(wb: Workbook, styles: dict[str, Any]) -> Worksheet:
    """Create the Identity sheet with MIA metadata."""
    ws = wb.active
    ws.title = "Identity"
    ws.sheet_properties.tabColor = "1B1B2F"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 60

    # Title block
    for row in range(IDENTITY_TITLE_ROW, IDENTITY_AUTHOR_ROW + 1):
        _fill_row_dark(ws, row, styles, range(1, 4))

    ws.merge_cells(f"B{IDENTITY_TITLE_ROW}:C{IDENTITY_TITLE_ROW}")
    cell = ws.cell(row=IDENTITY_TITLE_ROW, column=2)
    cell.value = "MIA - Map of Intent and Action"
    cell.font = styles["title_font"]
    cell.fill = styles["fill_dark"]
    cell.alignment = styles["align_left"]

    ws.merge_cells(f"B{IDENTITY_SUBTITLE_ROW}:C{IDENTITY_SUBTITLE_ROW}")
    cell = ws.cell(row=IDENTITY_SUBTITLE_ROW, column=2)
    cell.value = "A framework for mapping user intents into structured actions"
    cell.font = styles["subtitle_font"]
    cell.fill = styles["fill_dark"]
    cell.alignment = styles["align_left"]

    ws.merge_cells(f"B{IDENTITY_AUTHOR_ROW}:C{IDENTITY_AUTHOR_ROW}")
    cell = ws.cell(row=IDENTITY_AUTHOR_ROW, column=2)
    cell.value = AUTHOR_LINE
    cell.font = styles["author_font"]
    cell.fill = styles["fill_dark"]
    cell.alignment = styles["align_left"]

    ws.row_dimensions[IDENTITY_TITLE_ROW].height = 40
    ws.row_dimensions[IDENTITY_SUBTITLE_ROW].height = 25
    ws.row_dimensions[IDENTITY_AUTHOR_ROW].height = 22

    # Identity fields
    for i, field in enumerate(IDENTITY_FIELDS):
        row = IDENTITY_FIELDS_START_ROW + i
        ws.row_dimensions[row].height = 35

        cell_field = ws.cell(row=row, column=IDENTITY_FIELD_COL)
        cell_field.value = field
        cell_field.font = styles["field_font"]
        cell_field.fill = styles["fill_field"]
        cell_field.border = styles["border"]
        cell_field.alignment = styles["align_left"]

        cell_value = ws.cell(row=row, column=IDENTITY_VALUE_COL)
        cell_value.value = ""
        cell_value.font = styles["value_font"]
        cell_value.fill = styles["fill_value"]
        cell_value.border = styles["border"]
        cell_value.alignment = styles["align_left"]

    # Instructions
    row_inst = IDENTITY_FIELDS_START_ROW + len(IDENTITY_FIELDS) + 2
    ws.merge_cells(f"B{row_inst}:C{row_inst}")
    cell = ws.cell(row=row_inst, column=2)
    cell.value = "How to fill"
    cell.font = styles["instruction_title_font"]
    cell.alignment = styles["align_left"]

    for j, text in enumerate(INSTRUCTIONS):
        row_txt = row_inst + 1 + j
        ws.merge_cells(f"B{row_txt}:C{row_txt}")
        cell = ws.cell(row=row_txt, column=2)
        cell.value = text
        cell.font = styles["instruction_font"]
        cell.alignment = styles["align_left"]

    return ws


def create_map_sheet(wb: Workbook, styles: dict[str, Any]) -> Worksheet:
    """Create the Map sheet with the core table and examples."""
    ws = wb.create_sheet("Map")
    ws.sheet_properties.tabColor = "162447"

    widths = {"A": 5, "B": 40, "C": 45, "D": 45, "E": 55}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    col_range = range(1, MAP_LAST_COL + 1)

    # Title block
    for row in (MAP_TITLE_ROW, MAP_SUBTITLE_ROW, MAP_AUTHOR_ROW):
        _fill_row_dark(ws, row, styles, col_range)

    ws.merge_cells(f"B{MAP_TITLE_ROW}:E{MAP_TITLE_ROW}")
    cell = ws.cell(row=MAP_TITLE_ROW, column=2)
    cell.value = "Map of Intents"
    cell.font = styles["title_font"]
    cell.fill = styles["fill_dark"]
    cell.alignment = styles["align_left"]

    ws.merge_cells(f"B{MAP_SUBTITLE_ROW}:E{MAP_SUBTITLE_ROW}")
    cell = ws.cell(row=MAP_SUBTITLE_ROW, column=2)
    cell.value = "Question  >  Intent  >  Operation  >  Action"
    cell.font = styles["subtitle_font"]
    cell.fill = styles["fill_dark"]
    cell.alignment = styles["align_left"]

    ws.merge_cells(f"B{MAP_AUTHOR_ROW}:E{MAP_AUTHOR_ROW}")
    cell = ws.cell(row=MAP_AUTHOR_ROW, column=2)
    cell.value = AUTHOR_LINE
    cell.font = styles["author_font"]
    cell.fill = styles["fill_dark"]
    cell.alignment = styles["align_left"]

    ws.row_dimensions[MAP_TITLE_ROW].height = 40
    ws.row_dimensions[MAP_SUBTITLE_ROW].height = 25
    ws.row_dimensions[MAP_AUTHOR_ROW].height = 22

    # Table headers
    ws.row_dimensions[MAP_HEADER_ROW].height = 20
    ws.row_dimensions[MAP_DESC_ROW].height = 20

    for i, (header, desc) in enumerate(zip(MAP_HEADERS, MAP_DESCRIPTIONS)):
        col = MAP_FIRST_COL + i

        cell = ws.cell(row=MAP_HEADER_ROW, column=col)
        cell.value = header
        cell.font = styles["header_font"]
        cell.fill = styles["fill_header"]
        cell.border = styles["border"]
        cell.alignment = styles["align_center"]

        cell_desc = ws.cell(row=MAP_DESC_ROW, column=col)
        cell_desc.value = desc
        cell_desc.font = styles["desc_font"]
        cell_desc.fill = styles["fill_desc"]
        cell_desc.border = styles["border"]
        cell_desc.alignment = styles["align_center"]

    # Examples
    for i, example in enumerate(EXAMPLES):
        row = MAP_DATA_START_ROW + i
        ws.row_dimensions[row].height = 45
        fill = styles["fill_row_even"] if i % 2 == 0 else styles["fill_row_odd"]

        for j, value in enumerate(example):
            col = MAP_FIRST_COL + j
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.font = styles["value_font"]
            cell.fill = fill
            cell.border = styles["border"]
            cell.alignment = styles["align_left"]

    # Empty rows for filling
    for i in range(EMPTY_ROWS_COUNT):
        row = MAP_DATA_START_ROW + len(EXAMPLES) + i
        ws.row_dimensions[row].height = 35
        fill = styles["fill_row_even"] if (len(EXAMPLES) + i) % 2 == 0 else styles["fill_row_odd"]

        for col in range(MAP_FIRST_COL, MAP_LAST_COL + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = styles["border"]
            cell.alignment = styles["align_left"]
            cell.font = styles["value_font"]

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def generate(output: str = "MIA_Template.xlsx") -> str:
    """Generate the MIA template and return the output path."""
    wb = openpyxl.Workbook()
    styles = create_styles()

    create_identity_sheet(wb, styles)
    create_map_sheet(wb, styles)

    try:
        wb.save(output)
    except PermissionError:
        msg = f"Cannot write '{output}' — the file is open in another program. Close it and try again."
        logger.error(msg)
        raise MiaError(msg)
    except OSError as exc:
        msg = f"Failed to save '{output}': {exc}"
        logger.error(msg)
        raise MiaError(msg)

    logger.info("Template created: %s", output)
    return output


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(message)s",
    )

    parser = argparse.ArgumentParser(
        description="MIA Framework — Generate an Excel template",
    )
    parser.add_argument(
        "-o", "--output",
        default="MIA_Template.xlsx",
        help="Output file path (default: MIA_Template.xlsx)",
    )
    args = parser.parse_args()

    try:
        generate(args.output)
    except MiaError:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
