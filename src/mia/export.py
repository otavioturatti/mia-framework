"""
MIA Framework - Export
Map of Intent and Action

Reads a MIA file (.xlsx) and exports each row from the "Map" sheet
as a structured JSON document ready for ingestion into vector databases,
RAG pipelines, or any downstream system.

This module handles serialization only — embedding, indexing, retrieval,
and generation are the responsibility of your pipeline.

Usage:
    python export.py MIA_Template.xlsx
    python export.py MIA_Template.xlsx --output custom_name.json

Output:
    <input_stem>_export.json (default)
"""

from __future__ import annotations

import argparse
import json
import logging
import re
from pathlib import Path
from typing import Any

import openpyxl

from mia.exceptions import MiaError
from mia.generate_template import (
    IDENTITY_FIELD_COL,
    IDENTITY_FIELDS,
    IDENTITY_FIELDS_START_ROW,
    IDENTITY_VALUE_COL,
    MAP_DATA_START_ROW,
    MAP_FIRST_COL,
    MAP_HEADER_ROW,
    MAP_HEADERS,
    MAP_LAST_COL,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

IDENTITY_SHEET_NAME = "Identity"
MAP_SHEET_NAME = "Map"

# Explicit mapping between Excel labels and JSON keys.
# Positional correspondence is eliminated — each pair is self-documenting.
IDENTITY_FIELD_MAP: list[tuple[str, str]] = [
    ("Name", "name"),
    ("Domain", "domain"),
    ("Data Source", "data_source"),
    ("Agent", "agent"),
    ("Version", "version"),
    ("Description", "description"),
]

logger = logging.getLogger("mia.export")


# ---------------------------------------------------------------------------
# Readers
# ---------------------------------------------------------------------------

def _validate_identity_fields(ws: Any) -> None:
    """Verify that the Identity sheet fields match the expected structure."""
    expected_labels = [label for label, _key in IDENTITY_FIELD_MAP]
    actual = [
        ws.cell(row=IDENTITY_FIELDS_START_ROW + i, column=IDENTITY_FIELD_COL).value
        for i in range(len(IDENTITY_FIELD_MAP))
    ]
    if actual != expected_labels:
        msg = (
            f"Identity sheet fields don't match expected structure. "
            f"Expected {expected_labels}, found {actual}. "
            f"The template may have been modified."
        )
        logger.error(msg)
        raise MiaError(msg)


def read_identity(ws: Any) -> dict[str, str]:
    """Read metadata from the Identity sheet."""
    _validate_identity_fields(ws)
    identity: dict[str, str] = {}

    for i, (label, key) in enumerate(IDENTITY_FIELD_MAP):
        try:
            value = ws.cell(row=IDENTITY_FIELDS_START_ROW + i, column=IDENTITY_VALUE_COL).value
        except Exception as exc:
            logger.warning("Could not read identity field '%s' (%s): %s", key, label, exc)
            value = None
        identity[key] = str(value) if value else ""

    return identity


def _validate_map_headers(ws: Any) -> None:
    """Verify that the Map sheet headers match the expected structure."""
    actual = [
        ws.cell(row=MAP_HEADER_ROW, column=MAP_FIRST_COL + i).value
        for i in range(len(MAP_HEADERS))
    ]
    if actual != MAP_HEADERS:
        msg = (
            f"Map sheet headers don't match expected structure. "
            f"Expected {MAP_HEADERS}, found {actual}. "
            f"The template may have been modified."
        )
        logger.error(msg)
        raise MiaError(msg)


def read_map(ws: Any) -> list[dict[str, str]]:
    """Read rows from the Map sheet, skipping headers and empty rows."""
    _validate_map_headers(ws)
    rows: list[dict[str, str]] = []

    for row in ws.iter_rows(min_row=MAP_DATA_START_ROW, min_col=MAP_FIRST_COL, max_col=MAP_LAST_COL):
        values = [cell.value for cell in row]

        if not any(values):
            continue

        entry = {
            "question": str(values[0]) if values[0] else "",
            "intent": str(values[1]) if values[1] else "",
            "operation": str(values[2]) if values[2] else "",
            "action": str(values[3]) if values[3] else "",
        }

        if not entry["question"]:
            logger.warning(
                "Row %d has no Question — skipping.",
                row[0].row,
            )
            continue

        rows.append(entry)

    return rows


# ---------------------------------------------------------------------------
# Transformer
# ---------------------------------------------------------------------------

def export_documents(
    identity: dict[str, str],
    rows: list[dict[str, str]],
) -> list[dict[str, Any]]:
    """Export each row as a structured document for downstream consumption."""
    documents: list[dict[str, Any]] = []
    mia_id_prefix = re.sub(r"[^a-z0-9]+", "_", identity.get("name", "mia").lower()).strip("_")

    for i, row in enumerate(rows):
        # Primary embedding: Question + Intent separated by " — ".
        # The separator gives transformer-based models a structural signal
        # that these are two distinct semantic units (user phrasing vs. domain meaning).
        # Including Intent helps disambiguate duplicate questions
        # (e.g., same question for different user profiles).
        primary = f"{row['question']} — {row['intent']}".strip(" —")

        # Secondary embedding: Intent + Operation for cross-encoder reranking.
        # This captures the "what" and "how" without the raw user phrasing,
        # enabling precise reranking after initial retrieval.
        secondary = f"{row['intent']} — {row['operation']}".strip(" —")

        document = {
            "id": f"{mia_id_prefix}_{i + 1:03d}",
            "content_for_embedding": primary,
            "content_for_reranking": secondary,
            "metadata": {
                "question": row["question"],
                "intent": row["intent"],
                "operation": row["operation"],
                "action": row["action"],
                "domain": identity.get("domain", ""),
                "data_source": identity.get("data_source", ""),
                "agent": identity.get("agent", ""),
                "version": identity.get("version", ""),
                "mia_name": identity.get("name", ""),
            },
        }
        documents.append(document)

    return documents


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(input_path: Path, output_path: Path | None = None) -> Path:
    """Execute the full transformation pipeline and return the output path."""
    if not input_path.exists():
        msg = f"File not found: {input_path}"
        logger.error(msg)
        raise MiaError(msg)

    if input_path.suffix.lower() not in (".xlsx", ".xlsm"):
        msg = f"Unsupported file format '{input_path.suffix}'. Expected .xlsx"
        logger.error(msg)
        raise MiaError(msg)

    try:
        wb = openpyxl.load_workbook(input_path, read_only=True)
    except Exception as exc:
        msg = f"Failed to open '{input_path}': {exc}"
        logger.error(msg)
        raise MiaError(msg)

    try:
        if IDENTITY_SHEET_NAME not in wb.sheetnames:
            msg = f"'{IDENTITY_SHEET_NAME}' sheet not found in the file."
            logger.error(msg)
            raise MiaError(msg)

        if MAP_SHEET_NAME not in wb.sheetnames:
            msg = f"'{MAP_SHEET_NAME}' sheet not found in the file."
            logger.error(msg)
            raise MiaError(msg)

        identity = read_identity(wb[IDENTITY_SHEET_NAME])
        rows = read_map(wb[MAP_SHEET_NAME])
    finally:
        wb.close()

    if not rows:
        msg = f"No valid rows found in the '{MAP_SHEET_NAME}' sheet."
        logger.error(msg)
        raise MiaError(msg)

    documents = export_documents(identity, rows)

    if output_path is None:
        output_path = Path(input_path.stem + "_export.json")

    try:
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(documents, f, ensure_ascii=False, indent=2)
    except OSError as exc:
        msg = f"Failed to write '{output_path}': {exc}"
        logger.error(msg)
        raise MiaError(msg)

    logger.info("Transformation complete!")
    logger.info("  Input:     %s", input_path)
    logger.info("  Output:    %s", output_path)
    logger.info("  Documents: %d", len(documents))
    logger.info("  Domain:    %s", identity.get("domain") or "N/A")
    logger.info("  Version:   %s", identity.get("version") or "N/A")

    return output_path


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(message)s",
    )

    parser = argparse.ArgumentParser(
        description="MIA Framework — Export an Excel MIA into structured JSON",
    )
    parser.add_argument(
        "input",
        help="Path to the MIA .xlsx file",
    )
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="Output JSON file path (default: <input>_export.json)",
    )
    args = parser.parse_args()

    output = Path(args.output) if args.output else None
    try:
        run(Path(args.input), output)
    except MiaError:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
